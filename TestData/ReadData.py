import openpyxl

wk = openpyxl.load_workbook("C://developement//TestData100.xlsx")


def fetch_no_of_rows(Sheetname):
    sh = wk[Sheetname]  # creating object of sheet
    return sh.max_row


def fetch_no_of_cell(Sheetname, row, cell):
    sh = wk[Sheetname]
    cell = sh.cell(int(row), int(cell))
    return cell.value


#print(fetch_no_of_rows("Sheet1"))
#print(fetch_no_of_cell("Sheet1", 1, 2))
